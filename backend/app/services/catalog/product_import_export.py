import csv
import io
import uuid
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.exceptions.base import BadRequestError
from app.models.catalog.category import Category
from app.models.catalog.collection import Collection
from app.models.catalog.manufacturer import Manufacturer
from app.models.catalog.product import Product
from app.models.catalog.product_type import ProductType
from app.models.catalog.unit import Unit
from app.models.enums import ProductStatus
from app.schemas.catalog.product import ProductCreate, ProductUpdate
from app.schemas.catalog.product_import_export import (
    ImportCommitResponse,
    ImportMode,
    ImportPreviewResponse,
    ImportRowError,
    ImportRowPreview,
)
from app.services.catalog.product import ProductService

_EXPORT_COLUMNS = (
    "sku",
    "barcode",
    "name",
    "description",
    "category",
    "manufacturer",
    "product_type",
    "collection",
    "unit",
    "price",
    "compare_at_price",
    "cost_price",
    "sale_price",
    "currency",
    "stock_quantity",
    "status",
    "is_featured",
    "weight_kg",
    "slug",
)


@dataclass
class _ReferenceCache:
    categories: dict[str, uuid.UUID] = field(default_factory=dict)
    manufacturers: dict[str, uuid.UUID] = field(default_factory=dict)
    units: dict[str, uuid.UUID] = field(default_factory=dict)
    product_types: dict[str, uuid.UUID] = field(default_factory=dict)
    collections: dict[str, uuid.UUID] = field(default_factory=dict)
    existing_skus: dict[str, uuid.UUID] = field(default_factory=dict)


class ProductImportExportService:
    """Excel/CSV import & export for the product catalog. Import supports
    a dry-run preview (nothing written), duplicate/conflict detection by
    SKU, and three modes: full create-or-update, price-only, and
    stock-only. Designed for admin-driven batches (hundreds to low
    thousands of rows per file) rather than a background job queue, which
    doesn't exist in this project yet — see CATALOG.md."""

    def __init__(self, session: AsyncSession, product_service: ProductService) -> None:
        self.session = session
        self.product_service = product_service

    # ------------------------------------------------------------------
    # Parsing
    # ------------------------------------------------------------------

    def parse_file(self, filename: str, content: bytes) -> list[dict[str, str]]:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return self._parse_csv(content)
        if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
            return self._parse_xlsx(content)
        raise BadRequestError(key="errors.import_unsupported_format")

    def _parse_csv(self, content: bytes) -> list[dict[str, str]]:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [
            {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
            for row in reader
        ]

    def _parse_xlsx(self, content: bytes) -> list[dict[str, str]]:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except InvalidFileException as exc:
            raise BadRequestError(key="errors.import_unsupported_format") from exc

        sheet = workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(h or "").strip().lower() for h in next(rows_iter)]
        except StopIteration:
            return []

        rows: list[dict[str, str]] = []
        for raw_row in rows_iter:
            if all(cell is None for cell in raw_row):
                continue
            row = {
                headers[i]: ("" if cell is None else str(cell)).strip()
                for i, cell in enumerate(raw_row)
                if i < len(headers)
            }
            rows.append(row)
        return rows

    # ------------------------------------------------------------------
    # Reference resolution
    # ------------------------------------------------------------------

    async def _build_reference_cache(self) -> _ReferenceCache:
        cache = _ReferenceCache()

        for row in (await self.session.execute(select(Category.id, Category.slug))):
            cache.categories[row.slug] = row.id
        for row in (await self.session.execute(select(Manufacturer.id, Manufacturer.slug))):
            cache.manufacturers[row.slug] = row.id
        for row in (await self.session.execute(select(Unit.id, Unit.symbol))):
            cache.units[row.symbol] = row.id
        for row in (await self.session.execute(select(ProductType.id, ProductType.code))):
            cache.product_types[row.code] = row.id
        for row in (await self.session.execute(select(Collection.id, Collection.code))):
            cache.collections[row.code] = row.id
        for row in (
            await self.session.execute(
                select(Product.id, Product.sku).where(Product.deleted_at.is_(None))
            )
        ):
            cache.existing_skus[row.sku] = row.id

        return cache

    # ------------------------------------------------------------------
    # Validation (shared by preview and commit)
    # ------------------------------------------------------------------

    def _validate_row(
        self,
        row_number: int,
        row: dict[str, str],
        mode: ImportMode,
        cache: _ReferenceCache,
        seen_skus: set[str],
    ) -> ImportRowPreview:
        errors: list[str] = []
        sku = row.get("sku", "").strip()
        if not sku:
            errors.append("sku is required")
            return ImportRowPreview(row_number=row_number, action="skip", errors=errors, data=row)

        if sku in seen_skus:
            errors.append(f"duplicate sku '{sku}' within the file")
        seen_skus.add(sku)

        is_existing = sku in cache.existing_skus
        data: dict[str, Any] = {"sku": sku}

        def _decimal(field_name: str, required: bool = False) -> Decimal | None:
            raw = row.get(field_name, "").strip()
            if not raw:
                if required:
                    errors.append(f"{field_name} is required")
                return None
            try:
                return Decimal(raw)
            except InvalidOperation:
                errors.append(f"{field_name} must be a valid number")
                return None

        if mode == "price_only":
            price = _decimal("price", required=True)
            if not is_existing:
                errors.append(
                    f"sku '{sku}' does not exist — price-only mode cannot create products"
                )
            data["price"] = price
            action: Any = "update" if is_existing and not errors else "skip"
            return ImportRowPreview(
                row_number=row_number, sku=sku, action=action, errors=errors, data=data
            )

        if mode == "stock_only":
            stock_raw = row.get("stock_quantity", "").strip()
            try:
                stock_quantity = int(stock_raw) if stock_raw else None
            except ValueError:
                stock_quantity = None
                errors.append("stock_quantity must be a valid integer")
            if stock_quantity is None and not errors:
                errors.append("stock_quantity is required")
            if not is_existing:
                errors.append(
                    f"sku '{sku}' does not exist — stock-only mode cannot create products"
                )
            data["stock_quantity"] = stock_quantity
            action = "update" if is_existing and not errors else "skip"
            return ImportRowPreview(
                row_number=row_number, sku=sku, action=action, errors=errors, data=data
            )

        # full mode
        name = row.get("name", "").strip()
        if not name:
            errors.append("name is required")
        data["name"] = name

        slug = row.get("slug", "").strip() or sku.lower().replace(" ", "-")
        data["slug"] = slug

        category_slug = row.get("category", "").strip()
        category_id = cache.categories.get(category_slug)
        if not category_id:
            errors.append(f"unknown category '{category_slug}'")
        data["category_id"] = category_id

        unit_symbol = row.get("unit", "").strip()
        unit_id = cache.units.get(unit_symbol)
        if not unit_id:
            errors.append(f"unknown unit '{unit_symbol}'")
        data["unit_id"] = unit_id

        product_type_code = row.get("product_type", "").strip()
        product_type_id = cache.product_types.get(product_type_code)
        if not product_type_id:
            errors.append(f"unknown product_type '{product_type_code}'")
        data["product_type_id"] = product_type_id

        manufacturer_slug = row.get("manufacturer", "").strip()
        if manufacturer_slug:
            manufacturer_id = cache.manufacturers.get(manufacturer_slug)
            if not manufacturer_id:
                errors.append(f"unknown manufacturer '{manufacturer_slug}'")
            data["manufacturer_id"] = manufacturer_id

        collection_code = row.get("collection", "").strip()
        if collection_code:
            collection_id = cache.collections.get(collection_code)
            if not collection_id:
                errors.append(f"unknown collection '{collection_code}'")
            data["collection_id"] = collection_id

        data["price"] = _decimal("price", required=True)
        data["compare_at_price"] = _decimal("compare_at_price")
        data["cost_price"] = _decimal("cost_price")
        data["sale_price"] = _decimal("sale_price")

        stock_raw = row.get("stock_quantity", "").strip()
        if stock_raw:
            try:
                data["stock_quantity"] = int(stock_raw)
            except ValueError:
                errors.append("stock_quantity must be a valid integer")

        weight_raw = row.get("weight_kg", "").strip()
        if weight_raw:
            data["weight_kg"] = _decimal("weight_kg")

        status_raw = row.get("status", "").strip().upper()
        if status_raw:
            try:
                data["status"] = ProductStatus[status_raw]
            except KeyError:
                errors.append(f"unknown status '{status_raw}'")

        is_featured_raw = row.get("is_featured", "").strip().lower()
        if is_featured_raw:
            data["is_featured"] = is_featured_raw in ("1", "true", "yes")

        barcode = row.get("barcode", "").strip()
        if barcode:
            data["barcode"] = barcode

        description = row.get("description", "").strip()
        if description:
            data["description"] = description

        action = "update" if is_existing else "create"
        if errors:
            action = "skip"
        return ImportRowPreview(
            row_number=row_number, sku=sku, action=action, errors=errors, data=data
        )

    # ------------------------------------------------------------------
    # Preview (dry run)
    # ------------------------------------------------------------------

    async def preview_import(
        self, rows: list[dict[str, str]], mode: ImportMode
    ) -> ImportPreviewResponse:
        cache = await self._build_reference_cache()
        seen_skus: set[str] = set()
        results = [
            self._validate_row(i + 1, row, mode, cache, seen_skus) for i, row in enumerate(rows)
        ]
        duplicate_skus = sorted(
            {r.sku for r in results if r.sku and r.errors and "duplicate" in " ".join(r.errors)}
        )
        return ImportPreviewResponse(
            total_rows=len(results),
            valid_rows=sum(1 for r in results if r.action != "skip"),
            invalid_rows=sum(1 for r in results if r.action == "skip"),
            create_count=sum(1 for r in results if r.action == "create"),
            update_count=sum(1 for r in results if r.action == "update"),
            duplicate_skus_in_file=duplicate_skus,
            rows=results,
        )

    # ------------------------------------------------------------------
    # Commit
    # ------------------------------------------------------------------

    async def commit_import(
        self,
        rows: list[dict[str, str]],
        mode: ImportMode,
        *,
        atomic: bool = False,
        actor_id: uuid.UUID | None = None,
    ) -> ImportCommitResponse:
        cache = await self._build_reference_cache()
        seen_skus: set[str] = set()
        validated = [
            self._validate_row(i + 1, row, mode, cache, seen_skus) for i, row in enumerate(rows)
        ]

        if atomic and any(r.action == "skip" for r in validated):
            errors = [
                ImportRowError(row_number=r.row_number, sku=r.sku, message="; ".join(r.errors))
                for r in validated
                if r.action == "skip"
            ]
            return ImportCommitResponse(
                created_count=0, updated_count=0, skipped_count=len(errors), errors=errors
            )

        created_count = 0
        updated_count = 0
        errors = []

        for result in validated:
            if result.action == "skip":
                errors.append(
                    ImportRowError(
                        row_number=result.row_number,
                        sku=result.sku,
                        message="; ".join(result.errors) or "skipped",
                    )
                )
                continue

            try:
                if mode == "price_only":
                    product_id = cache.existing_skus[result.sku]  # type: ignore[index]
                    await self.product_service.update(
                        product_id, ProductUpdate(price=result.data["price"]), actor_id
                    )
                    updated_count += 1
                elif mode == "stock_only":
                    product_id = cache.existing_skus[result.sku]  # type: ignore[index]
                    await self.product_service.update(
                        product_id,
                        ProductUpdate(stock_quantity=result.data["stock_quantity"]),
                        actor_id,
                    )
                    updated_count += 1
                elif result.action == "create":
                    await self.product_service.create(ProductCreate(**result.data), actor_id)
                    created_count += 1
                else:
                    product_id = cache.existing_skus[result.sku]  # type: ignore[index]
                    payload = {k: v for k, v in result.data.items() if k != "sku"}
                    await self.product_service.update(
                        product_id, ProductUpdate(**payload), actor_id
                    )
                    updated_count += 1
            except Exception as exc:  # noqa: BLE001 — one bad row must not sink the whole batch
                errors.append(
                    ImportRowError(row_number=result.row_number, sku=result.sku, message=str(exc))
                )

        return ImportCommitResponse(
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=len(errors),
            errors=errors,
        )

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    async def export_products(self, products: list[Product], fmt: str) -> tuple[bytes, str]:
        rows = []
        for product in products:
            rows.append(
                {
                    "sku": product.sku,
                    "barcode": product.barcode or "",
                    "name": product.name,
                    "description": product.description or "",
                    "category": product.category.slug if product.category else "",
                    "manufacturer": product.manufacturer.slug if product.manufacturer else "",
                    "product_type": product.product_type.code if product.product_type else "",
                    "collection": product.collection.code if product.collection else "",
                    "unit": product.unit.symbol if product.unit else "",
                    "price": str(product.price),
                    "compare_at_price": str(product.compare_at_price or ""),
                    "cost_price": str(product.cost_price or ""),
                    "sale_price": str(product.sale_price or ""),
                    "currency": product.currency,
                    "stock_quantity": product.stock_quantity,
                    "status": product.status.value,
                    "is_featured": product.is_featured,
                    "weight_kg": str(product.weight_kg or ""),
                    "slug": product.slug,
                }
            )

        if fmt == "csv":
            buffer = io.StringIO()
            writer = csv.DictWriter(buffer, fieldnames=list(_EXPORT_COLUMNS))
            writer.writeheader()
            writer.writerows(rows)
            return buffer.getvalue().encode("utf-8-sig"), "text/csv"

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.append(list(_EXPORT_COLUMNS))
        for row in rows:
            sheet.append([row[col] for col in _EXPORT_COLUMNS])
        out = io.BytesIO()
        workbook.save(out)
        return (
            out.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
